"""Advanced Excel tools with formula support."""
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import io
from typing import Dict, List, Optional, Any


class ExcelTools:
    """Advanced Excel manipulation with formulas, formatting, and cross-sheet operations."""

    @staticmethod
    def load_excel(file, header_row: int = None) -> Dict:
        """
        Load Excel file with intelligent header detection and cleaning.
        Handles real-world dirty datasets.

        Args:
            file: File object or path
            header_row: Specific row to use as header (auto-detect if None)

        Returns:
            Dict with dataframe and sheet info
        """
        # First, read raw to detect structure
        try:
            df_raw = pd.read_excel(file, header=None, nrows=20)
        except:
            file.seek(0)
            df_raw = pd.read_excel(file, header=None, nrows=20)
        
        # Detect actual header row if not specified
        if header_row is None:
            header_row = ExcelTools._detect_actual_header(df_raw)
        
        # Reset file pointer
        try:
            file.seek(0)
        except:
            pass
        
        # Load with detected header
        df = pd.read_excel(file, header=header_row)
        
        # Clean the dataframe
        df = ExcelTools._clean_loaded_data(df)

        return {
            'dataframe': df,
            'active_sheet': 'Sheet1',
            'header_row': header_row,
            'cleaned': True
        }
    
    @staticmethod
    def _detect_actual_header(df_raw: pd.DataFrame) -> int:
        """
        Detect which row contains actual headers.
        Handles cases where first few rows are empty or metadata.
        """
        for idx in range(min(10, len(df_raw))):
            row = df_raw.iloc[idx]
            
            # Skip completely empty rows
            if row.isna().all():
                continue
            
            # Count non-null values
            non_null = row.notna().sum()
            
            # Count string values (headers are usually strings)
            string_count = sum(1 for val in row if isinstance(val, str) and len(str(val).strip()) > 0)
            
            # If row has >50% non-null AND >50% strings, likely a header
            if non_null > len(row) * 0.5 and string_count > len(row) * 0.5:
                # Check if values look like headers (not numeric, not dates)
                numeric_count = sum(1 for val in row if isinstance(val, (int, float)) and not isinstance(val, bool))
                if numeric_count < len(row) * 0.3:  # Less than 30% numeric
                    return idx
        
        return 0  # Default to first row
    
    @staticmethod
    def _clean_loaded_data(df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean loaded dataframe from real-world issues.
        """
        # Remove completely empty rows and columns
        df = df.dropna(how='all').dropna(axis=1, how='all')
        
        # Clean column names
        import re
        cleaned_cols = []
        for i, col in enumerate(df.columns):
            col_str = str(col).strip()
            
            # Handle Unnamed columns
            if 'Unnamed' in col_str or col_str == '' or col_str == 'nan':
                col_str = f'Column_{i}'
            
            # Remove special characters and extra spaces
            col_str = re.sub(r'[\n\r\t]+', ' ', col_str)
            col_str = re.sub(r'\s+', ' ', col_str)
            col_str = col_str.strip()
            
            cleaned_cols.append(col_str if col_str else f'Column_{i}')
        
        df.columns = cleaned_cols
        
        # Fix duplicate column names
        df = ExcelTools._fix_duplicate_column_names(df)
        
        # Reset index
        df = df.reset_index(drop=True)
        
        return df
    
    @staticmethod
    def _fix_duplicate_column_names(df: pd.DataFrame) -> pd.DataFrame:
        """Fix duplicate column names."""
        cols = pd.Series(df.columns)
        for dup in cols[cols.duplicated()].unique():
            dup_indices = cols[cols == dup].index.tolist()
            for i, idx in enumerate(dup_indices):
                if i > 0:
                    cols.iloc[idx] = f'{dup}_{i}'
        df.columns = cols
        return df

    @staticmethod
    def export_to_excel(df: pd.DataFrame, filename: str) -> io.BytesIO:
        """Export dataframe to Excel."""
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Data')
        output.seek(0)
        return output

    @staticmethod
    def apply_conditional_formatting(df: pd.DataFrame, rules: List[Dict]) -> io.BytesIO:
        """
        Apply conditional formatting to Excel file.

        Args:
            df: DataFrame to format
            rules: List of formatting rules
                - type: 'highlight_max', 'highlight_min', 'highlight_value', 'color_scale'
                - columns: list of column names
                - color: color code (e.g., 'FF0000' for red)
                - value: comparison value

        Returns:
            BytesIO with formatted Excel
        """
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Data')
            workbook = writer.book
            worksheet = writer.sheets['Data']

            # Apply rules
            for rule in rules:
                rule_type = rule.get('type', '')
                columns = rule.get('columns', [])
                color = rule.get('color', 'FFFF00')  # Yellow default

                for col_name in columns:
                    if col_name not in df.columns:
                        continue

                    col_idx = list(df.columns).index(col_name) + 2  # +1 for header, +1 for Excel
                    col_letter = get_column_letter(col_idx)

                    if rule_type == 'highlight_max':
                        # Highlight maximum value in column
                        max_val = df[col_name].max()
                        for row_idx, val in enumerate(df[col_name], start=2):
                            if val == max_val:
                                cell = worksheet[f'{col_letter}{row_idx}']
                                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                                cell.font = Font(bold=True)

                    elif rule_type == 'highlight_min':
                        # Highlight minimum value
                        min_val = df[col_name].min()
                        for row_idx, val in enumerate(df[col_name], start=2):
                            if val == min_val:
                                cell = worksheet[f'{col_letter}{row_idx}']
                                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

                    elif rule_type == 'highlight_value':
                        # Highlight specific values
                        target_val = rule.get('value')
                        for row_idx, val in enumerate(df[col_name], start=2):
                            if val == target_val:
                                cell = worksheet[f'{col_letter}{row_idx}']
                                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        output.seek(0)
        return output

    @staticmethod
    def create_vlookup_sheet(df: pd.DataFrame, lookup_col: str, return_col: str, 
                            sheet_name: str = 'Lookup') -> io.BytesIO:
        """
        Create a new sheet with VLOOKUP formulas.

        Args:
            df: Main dataframe
            lookup_col: Column to lookup
            return_col: Column to return
            sheet_name: Name of new sheet

        Returns:
            BytesIO with workbook containing VLOOKUP
        """
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Data')
            workbook = writer.book

            # Create lookup sheet
            lookup_sheet = workbook.create_sheet(sheet_name)

            # Add headers
            lookup_sheet['A1'] = lookup_col
            lookup_sheet['B1'] = f'Matched_{return_col}'
            lookup_sheet['A1'].font = Font(bold=True)
            lookup_sheet['B1'].font = Font(bold=True)

            # Get unique values from lookup column
            unique_vals = df[lookup_col].unique()

            for idx, val in enumerate(unique_vals, start=2):
                lookup_sheet[f'A{idx}'] = val

                # Create VLOOKUP formula
                vlookup_formula = f'=IFERROR(VLOOKUP(A{idx},Data!$A$1:$Z$1000,COLUMN(Data!${return_col}$1),FALSE),"")'
                lookup_sheet[f'B{idx}'] = vlookup_formula

        output.seek(0)
        return output

    @staticmethod
    def filter_by_criteria(df: pd.DataFrame, criteria: Dict) -> pd.DataFrame:
        """
        Filter dataframe by multiple criteria.

        Args:
            df: DataFrame to filter
            criteria: Dict of column -> value/condition pairs
                Examples:
                - {'Score': ('highest', 1)} -> Top 1 highest scores
                - {'Score': ('top', 10)} -> Top 10 scores
                - {'City': 'Delhi'} -> Exact match
                - {'Age': ('>', 25)} -> Greater than 25

        Returns:
            Filtered DataFrame
        """
        result = df.copy()

        for col, condition in criteria.items():
            if col not in result.columns:
                continue

            if isinstance(condition, tuple):
                op, val = condition

                if op == 'highest':
                    # Get top N highest values
                    result = result.nlargest(val, col)
                elif op == 'lowest':
                    result = result.nsmallest(val, col)
                elif op == 'top':
                    result = result.nlargest(val, col)
                elif op == 'bottom':
                    result = result.nsmallest(val, col)
                elif op == '>':
                    result = result[result[col] > val]
                elif op == '<':
                    result = result[result[col] < val]
                elif op == '>=':
                    result = result[result[col] >= val]
                elif op == '<=':
                    result = result[result[col] <= val]
                elif op == '==':
                    result = result[result[col] == val]
                elif op == '!=':
                    result = result[result[col] != val]
                elif op == 'contains':
                    result = result[result[col].astype(str).str.contains(str(val))]
            else:
                # Simple equality
                result = result[result[col] == condition]

        return result

    @staticmethod
    def create_summary_table(df: pd.DataFrame, group_by: str, 
                            aggregate: Dict) -> pd.DataFrame:
        """
        Create summary/pivot table with aggregations.

        Args:
            df: DataFrame
            group_by: Column to group by
            aggregate: Dict of column -> aggregation function
                - 'sum', 'mean', 'count', 'min', 'max', 'std'

        Returns:
            Aggregated DataFrame
        """
        agg_funcs = {}
        for col, func in aggregate.items():
            if col in df.columns and func in ['sum', 'mean', 'count', 'min', 'max', 'std']:
                agg_funcs[col] = func

        if not agg_funcs:
            return df

        return df.groupby(group_by).agg(agg_funcs).reset_index()

    @staticmethod
    def merge_sheets(df1: pd.DataFrame, df2: pd.DataFrame, on: str, 
                    how: str = 'inner') -> pd.DataFrame:
        """
        Merge two dataframes (cross-sheet verification).

        Args:
            df1: First dataframe
            df2: Second dataframe
            on: Column(s) to merge on
            how: 'inner', 'outer', 'left', 'right'

        Returns:
            Merged DataFrame
        """
        return pd.merge(df1, df2, on=on, how=how)

    @staticmethod
    def create_formulas_sheet(df: pd.DataFrame, formulas: Dict) -> io.BytesIO:
        """
        Create Excel sheet with custom formulas.

        Args:
            df: DataFrame
            formulas: Dict of sheet_name -> list of formulas
                Example:
                {
                    'Totals': [
                        ('A1', 'Total Orders'),
                        ('A2', '=SUM(Data!B:B)'),
                        ('A3', 'Average Price'),
                        ('A4', '=AVERAGE(Data!C:C)')
                    ]
                }

        Returns:
            BytesIO with Excel file
        """
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Data')
            workbook = writer.book

            # Add formula sheets
            for sheet_name, formula_list in formulas.items():
                formula_sheet = workbook.create_sheet(sheet_name)

                for cell_ref, value in formula_list:
                    cell = formula_sheet[cell_ref]
                    if isinstance(value, str) and value.startswith('='):
                        cell.value = value
                    else:
                        cell.value = value

                    if isinstance(value, str) and not value.startswith('='):
                        cell.font = Font(bold=True)

        output.seek(0)
        return output
